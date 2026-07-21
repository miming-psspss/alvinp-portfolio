# widgets/transaction_window.py
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import subprocess
import sys

class TransactionWindow:
    def __init__(self, parent, auth_system, db_manager, transaction_manager, on_back):
        self.parent = parent
        self.auth_system = auth_system
        self.db_manager = db_manager
        self.transaction_manager = transaction_manager
        self.on_back = on_back
        
        self.frame = ttk.Frame(parent)
        self.mode = "add"  # "add" or "view"
        self.create_widgets()
    
    def create_widgets(self):
        # Header
        header_frame = ttk.Frame(self.frame)
        header_frame.pack(fill='x', padx=20, pady=10)
        
        self.back_btn = ttk.Button(header_frame, text="← Back to Dashboard", 
                                  command=self.on_back)
        self.back_btn.pack(side='left')
        
        self.title_label = ttk.Label(header_frame, text="Add New Transaction", 
                                    font=('Arial', 16, 'bold'))
        self.title_label.pack(side='left', padx=20)
        
        # Main content
        self.main_frame = ttk.Frame(self.frame)
        self.main_frame.pack(expand=True, fill='both', padx=20, pady=10)
        
        self.create_add_form()
        self.create_view_section()
    
    def create_add_form(self):
        """Create the transaction addition form"""
        self.form_frame = ttk.LabelFrame(self.main_frame, text="Transaction Details", padding="20")
        
        # Date
        ttk.Label(self.form_frame, text="Date:").grid(row=0, column=0, sticky='w', pady=8)
        self.trans_date = ttk.Entry(self.form_frame, font=('Arial', 10))
        self.trans_date.insert(0, date.today().isoformat())
        self.trans_date.grid(row=0, column=1, sticky='ew', pady=8, padx=10)
        
        # Transaction Type
        ttk.Label(self.form_frame, text="Type:").grid(row=1, column=0, sticky='w', pady=8)
        self.trans_type = ttk.Combobox(self.form_frame, values=['Income', 'Expense'], 
                                      state='readonly', font=('Arial', 10))
        self.trans_type.set('Expense')
        self.trans_type.grid(row=1, column=1, sticky='ew', pady=8, padx=10)
        self.trans_type.bind('<<ComboboxSelected>>', self.update_categories)
        
        # Category
        ttk.Label(self.form_frame, text="Category:").grid(row=2, column=0, sticky='w', pady=8)
        self.category_combo = ttk.Combobox(self.form_frame, state='readonly', font=('Arial', 10))
        self.category_combo.grid(row=2, column=1, sticky='ew', pady=8, padx=10)
        
        # Amount
        ttk.Label(self.form_frame, text="Amount (₱):").grid(row=3, column=0, sticky='w', pady=8)
        self.amount_entry = ttk.Entry(self.form_frame, font=('Arial', 10))
        self.amount_entry.grid(row=3, column=1, sticky='ew', pady=8, padx=10)
        
        # Description
        ttk.Label(self.form_frame, text="Description:").grid(row=4, column=0, sticky='w', pady=8)
        self.desc_entry = ttk.Entry(self.form_frame, font=('Arial', 10))
        self.desc_entry.grid(row=4, column=1, sticky='ew', pady=8, padx=10)
        
        # Payee/Payer
        ttk.Label(self.form_frame, text="Payee/Payer:").grid(row=5, column=0, sticky='w', pady=8)
        self.payee_entry = ttk.Entry(self.form_frame, font=('Arial', 10))
        self.payee_entry.grid(row=5, column=1, sticky='ew', pady=8, padx=10)
        
        # Payment Method
        ttk.Label(self.form_frame, text="Payment Method:").grid(row=6, column=0, sticky='w', pady=8)
        self.payment_method = ttk.Combobox(self.form_frame, 
                                         values=['Cash', 'Check', 'Bank Transfer'], 
                                         state='readonly', font=('Arial', 10))
        self.payment_method.set('Cash')
        self.payment_method.grid(row=6, column=1, sticky='ew', pady=8, padx=10)
        
        # Check Number (only show when payment method is Check)
        self.check_label = ttk.Label(self.form_frame, text="Check Number:")
        self.check_entry = ttk.Entry(self.form_frame, font=('Arial', 10))
        self.payment_method.bind('<<ComboboxSelected>>', self.toggle_check_number)
        
        # Submit button
        self.submit_btn = ttk.Button(self.form_frame, text="Submit Transaction", 
                                    command=self.submit_transaction, width=20)
        self.submit_btn.grid(row=8, column=0, columnspan=2, pady=20)
        
        # Configure grid weights
        self.form_frame.columnconfigure(1, weight=1)
        
        # Initialize categories
        self.update_categories()
    
    def create_view_section(self):
        """Create the transaction viewing section"""
        self.view_frame = ttk.LabelFrame(self.main_frame, text="Recent Transactions", padding="10")
        
        # Filter frame
        filter_frame = ttk.Frame(self.view_frame)
        filter_frame.pack(fill='x', pady=5)
        
        ttk.Label(filter_frame, text="Filter:").pack(side='left')
        
        self.filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(filter_frame, text="All", variable=self.filter_var, value="all",
                       command=self.load_transactions).pack(side='left', padx=10)
        ttk.Radiobutton(filter_frame, text="Today", variable=self.filter_var, value="today",
                       command=self.load_transactions).pack(side='left', padx=10)
        ttk.Radiobutton(filter_frame, text="This Month", variable=self.filter_var, value="month",
                       command=self.load_transactions).pack(side='left', padx=10)
        
        # Action buttons frame
        action_frame = ttk.Frame(filter_frame)
        action_frame.pack(side='right')
        
        # Refresh button
        ttk.Button(action_frame, text="Refresh", 
                  command=self.load_transactions).pack(side='left', padx=5)
        
        # NEW: Print button
        self.print_btn = ttk.Button(action_frame, text="📄 Print Selected", 
                                   command=self.print_selected_transaction,
                                   state='disabled')
        self.print_btn.pack(side='left', padx=5)
        
        # Treeview for transactions
        columns = ('date', 'number', 'type', 'category', 'amount', 'status', 'description')
        self.transactions_tree = ttk.Treeview(self.view_frame, columns=columns, show='headings', height=12)
        
        # Define headings
        self.transactions_tree.heading('date', text='Date')
        self.transactions_tree.heading('number', text='Trans No')
        self.transactions_tree.heading('type', text='Type')
        self.transactions_tree.heading('category', text='Category')
        self.transactions_tree.heading('amount', text='Amount')
        self.transactions_tree.heading('status', text='Status')
        self.transactions_tree.heading('description', text='Description')
        
        # Configure columns
        self.transactions_tree.column('date', width=100)
        self.transactions_tree.column('number', width=120)
        self.transactions_tree.column('type', width=80)
        self.transactions_tree.column('category', width=120)
        self.transactions_tree.column('amount', width=100)
        self.transactions_tree.column('status', width=80)
        self.transactions_tree.column('description', width=200)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.view_frame, orient=tk.VERTICAL, command=self.transactions_tree.yview)
        self.transactions_tree.configure(yscrollcommand=scrollbar.set)
        
        self.transactions_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Bind selection event to enable/disable print button
        self.transactions_tree.bind('<<TreeviewSelect>>', self.on_treeview_select)
    
    def on_treeview_select(self, event):
        """Enable print button when a transaction is selected"""
        selection = self.transactions_tree.selection()
        if selection:
            self.print_btn.config(state='normal')
        else:
            self.print_btn.config(state='disabled')
    
    def toggle_check_number(self, event=None):
        """Show/hide check number field based on payment method"""
        if self.payment_method.get() == 'Check':
            self.check_label.grid(row=7, column=0, sticky='w', pady=8)
            self.check_entry.grid(row=7, column=1, sticky='ew', pady=8, padx=10)
        else:
            self.check_label.grid_remove()
            self.check_entry.grid_remove()
    
    def update_categories(self, event=None):
        """Update categories based on selected transaction type"""
        trans_type = self.trans_type.get().lower()
        with self.db_manager.get_connection() as conn:
            categories = conn.execute(
                "SELECT category_name FROM categories WHERE category_type = ?",
                (trans_type,)
            ).fetchall()
        
        category_names = [cat['category_name'] for cat in categories]
        self.category_combo['values'] = category_names
        if category_names:
            self.category_combo.set(category_names[0])
    
    def submit_transaction(self):
        """Submit a new transaction"""
        try:
            # Validate required fields
            if not all([self.amount_entry.get(), self.desc_entry.get()]):
                messagebox.showerror("Error", "Please fill in all required fields")
                return
            
            transaction_data = {
                'transaction_date': self.trans_date.get(),
                'category_type': self.trans_type.get().lower(),
                'category_name': self.category_combo.get(),
                'amount': float(self.amount_entry.get()),
                'description': self.desc_entry.get(),
                'payee_payer': self.payee_entry.get(),
                'payment_method': self.payment_method.get().lower(),
                'check_number': self.check_entry.get() if self.payment_method.get() == 'Check' else None
            }
            
            trans_number = self.transaction_manager.add_transaction(transaction_data)
            
            messagebox.showinfo("Success", 
                              f"Transaction added successfully!\nTransaction Number: {trans_number}")
            
            # Clear form
            self.clear_form()
            # Refresh transactions list
            self.load_transactions()
            
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add transaction: {str(e)}")
    
    def clear_form(self):
        """Clear the transaction form"""
        self.amount_entry.delete(0, tk.END)
        self.desc_entry.delete(0, tk.END)
        self.payee_entry.delete(0, tk.END)
        self.check_entry.delete(0, tk.END)
        self.trans_date.delete(0, tk.END)
        self.trans_date.insert(0, date.today().isoformat())
        self.desc_entry.focus()
    
    def load_transactions(self):
        """Load transactions into the treeview"""
        # Clear existing data
        for item in self.transactions_tree.get_children():
            self.transactions_tree.delete(item)
        
        # Get filter
        filter_type = self.filter_var.get()
        
        # Load transactions
        transactions = self.transaction_manager.get_transactions()
        
        # Apply filters
        from datetime import datetime, date
        today = date.today()
        
        filtered_transactions = []
        for trans in transactions:
            trans_date = datetime.strptime(trans['transaction_date'], '%Y-%m-%d').date()
            
            if filter_type == 'today' and trans_date != today:
                continue
            elif filter_type == 'month' and (trans_date.year != today.year or trans_date.month != today.month):
                continue
            
            filtered_transactions.append(trans)
        
        # Populate treeview
        for trans in filtered_transactions:
            trans_type = "INC" if trans['category_type'] == 'income' else "EXP"
            status = trans['status'].capitalize()
            
            # Color coding based on status
            tags = ()
            if status == 'Pending':
                tags = ('pending',)
            elif status == 'Approved':
                tags = ('approved',)
            elif status == 'Rejected':
                tags = ('rejected',)
            
            self.transactions_tree.insert('', 'end', values=(
                trans['transaction_date'],
                trans['transaction_number'],
                trans_type,
                trans['category_name'],
                f"₱{trans['amount']:,.2f}",
                status,
                trans['description']
            ), tags=tags)
        
        # Configure tag colors
        self.transactions_tree.tag_configure('pending', background='#fff3cd')
        self.transactions_tree.tag_configure('approved', background='#d1ecf1')
        self.transactions_tree.tag_configure('rejected', background='#f8d7da')
        
        # Disable print button after loading
        self.print_btn.config(state='disabled')
    
    def print_selected_transaction(self):
        """Generate Excel file for selected transaction and show print preview"""
        try:
            selection = self.transactions_tree.selection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a transaction to print")
                return
            
            # Get selected transaction data
            item = self.transactions_tree.item(selection[0])
            values = item['values']
            
            # Extract transaction data
            trans_data = {
                'date': values[0],
                'trans_no': values[1],
                'type': values[2],
                'category': values[3],
                'amount': values[4],
                'status': values[5],
                'description': values[6]
            }
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Transaction Receipt"
            
            # Define styles
            header_font = Font(name='Arial', size=12, bold=True)
            label_font = Font(name='Arial', size=10, bold=True)
            value_font = Font(name='Arial', size=10)
            footer_font = Font(name='Arial', size=8, italic=True)
            signature_font = Font(name='Arial', size=9)
            
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Calculate optimal column widths based on content
            header_text = "BARANGAY TRANSACTION RECEIPT"
            max_label_width = max(len(label) for label, _ in [
                ("Date:", trans_data['date']),
                ("Transaction No:", trans_data['trans_no']),
                ("Type:", trans_data['type']),
                ("Category:", trans_data['category']),
                ("Amount:", trans_data['amount']),
                ("Status:", trans_data['status']),
                ("Description:", trans_data['description'])
            ])
            
            # Set column widths dynamically
            ws.column_dimensions['A'].width = max_label_width + 2  # Label column
            ws.column_dimensions['B'].width = 25  # Value column
            
            # Header - automatically adjust to text length
            header_cols = 2  # Use both columns for header
            ws.merge_cells(f'A1:B1')
            ws['A1'] = header_text
            ws['A1'].font = header_font
            ws['A1'].alignment = center_align
            
            # Auto-adjust header row height if needed
            header_length = len(header_text)
            if header_length > 30:
                ws.row_dimensions[1].height = 25
            else:
                ws.row_dimensions[1].height = 20
            
            # Empty row for spacing
            ws.row_dimensions[2].height = 5
            
            # Transaction details
            details = [
                ("Date:", trans_data['date']),
                ("Transaction No:", trans_data['trans_no']),
                ("Type:", trans_data['type']),
                ("Category:", trans_data['category']),
                ("Amount:", trans_data['amount']),
                ("Status:", trans_data['status']),
                ("Description:", trans_data['description'])
            ]
            
            for i, (label, value) in enumerate(details, start=3):
                # Label cell
                ws[f'A{i}'] = label
                ws[f'A{i}'].font = label_font
                ws[f'A{i}'].alignment = left_align
                
                # Value cell
                ws[f'B{i}'] = value
                ws[f'B{i}'].font = value_font
                ws[f'B{i}'].alignment = left_align
            
            # Add signature section on the left side
            signature_start_row = len(details) + 4
            
            # Captain's Signature
            ws.merge_cells(f'A{signature_start_row}:B{signature_start_row}')
            ws[f'A{signature_start_row}'] = "_________________________"
            ws[f'A{signature_start_row}'].font = signature_font
            ws[f'A{signature_start_row}'].alignment = left_align
            
            ws.merge_cells(f'A{signature_start_row + 1}:B{signature_start_row + 1}')
            ws[f'A{signature_start_row + 1}'] = "Captain's Signature"
            ws[f'A{signature_start_row + 1}'].font = signature_font
            ws[f'A{signature_start_row + 1}'].alignment = left_align
            
            # Empty space between signatures
            ws.row_dimensions[signature_start_row + 2].height = 10
            
            # Treasurer's Signature
            ws.merge_cells(f'A{signature_start_row + 3}:B{signature_start_row + 3}')
            ws[f'A{signature_start_row + 3}'] = "_________________________"
            ws[f'A{signature_start_row + 3}'].font = signature_font
            ws[f'A{signature_start_row + 3}'].alignment = left_align
            
            ws.merge_cells(f'A{signature_start_row + 4}:B{signature_start_row + 4}')
            ws[f'A{signature_start_row + 4}'] = "Treasurer's Signature"
            ws[f'A{signature_start_row + 4}'].font = signature_font
            ws[f'A{signature_start_row + 4}'].alignment = left_align
            
            # Footer
            footer_row = signature_start_row + 6
            ws.merge_cells(f'A{footer_row}:B{footer_row}')
            ws[f'A{footer_row}'] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws[f'A{footer_row}'].font = footer_font
            ws[f'A{footer_row}'].alignment = center_align
            
            # Save the file
            filename = f"Transaction_{trans_data['trans_no']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            
            wb.save(filepath)
            
            # Show print preview dialog
            self.show_print_preview(filepath, trans_data)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate receipt: {str(e)}")

    def show_print_preview(self, filepath, trans_data):
        """Show a print preview dialog with options"""
        preview_window = tk.Toplevel(self.parent)
        preview_window.title("Print Preview - Transaction Receipt")
        preview_window.geometry("500x650")  # Increased height for signatures
        preview_window.transient(self.parent)
        preview_window.grab_set()
        
        # Center the preview window
        preview_window.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - preview_window.winfo_width()) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - preview_window.winfo_height()) // 2
        preview_window.geometry(f"+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(preview_window, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Preview title
        ttk.Label(main_frame, text="PRINT PREVIEW", 
                font=('Arial', 16, 'bold')).pack(pady=10)
        
        # Receipt preview frame
        receipt_frame = ttk.LabelFrame(main_frame, text="Receipt Preview", padding="15")
        receipt_frame.pack(fill='both', expand=True, pady=10)
        
        # Receipt content with signatures - FIXED BORDER ALIGNMENT
        receipt_content = f"""
    ╔══════════════════════════════════════╗
    ║      BARANGAY TRANSACTION RECEIPT    ║
    ╠══════════════════════════════════════╣
    ║ Date:           {trans_data['date']:<18}   ║
    ║ Transaction No: {trans_data['trans_no']:<18}   ║
    ║ Type:           {trans_data['type']:<18}   ║
    ║ Category:       {trans_data['category']:<18}   ║
    ║ Amount:         {trans_data['amount']:<18}   ║
    ║ Status:         {trans_data['status']:<18}   ║
    ║ Description:    {trans_data['description'][:18]:<18}   ║
    ╠══════════════════════════════════════╣
    ║ _________________________            ║
    ║ Captain's Signature                  ║
    ║                                      ║
    ║ _________________________            ║
    ║ Treasurer's Signature                ║
    ╠══════════════════════════════════════╣
    ║    Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}     ║
    ╚══════════════════════════════════════╝
    """
        
        # Preview text
        preview_text = tk.Text(receipt_frame, height=18, width=45,
                            font=('Courier New', 10), relief='solid', bd=1)
        preview_text.pack(pady=10, fill='both', expand=True)
        preview_text.insert('1.0', receipt_content)
        preview_text.config(state='disabled')  # Make it read-only
        
        # File info
        file_info = f"File: {os.path.basename(filepath)}\nLocation: {os.path.dirname(filepath)}"
        ttk.Label(main_frame, text=file_info, font=('Arial', 9), 
                foreground='blue').pack(pady=5)
        
        # Action buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        # Print button
        ttk.Button(button_frame, text="🖨️ Open File for Printing", 
                command=lambda: self.open_file_for_printing(filepath),
                width=20).pack(side='left', padx=10)
        
        # Open folder button
        ttk.Button(button_frame, text="📁 Open File Location", 
                command=lambda: self.open_file_location(filepath),
                width=18).pack(side='left', padx=10)
        
        # Close button
        ttk.Button(button_frame, text="Close Preview", 
                command=preview_window.destroy,
                width=15).pack(side='left', padx=10)

    def open_file_for_printing(self, filepath):
        """Open the Excel file for printing"""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(filepath)
            elif os.name == 'posix':  # macOS, Linux
                if sys.platform == "darwin":
                    subprocess.call(('open', filepath))
                else:
                    subprocess.call(('xdg-open', filepath))
            
            messagebox.showinfo("File Opened", 
                            "Excel file has been opened.\n\n"
                            "Please use Excel's print function (Ctrl+P) to print the receipt.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {str(e)}")

    def open_file_location(self, filepath):
        """Open the folder containing the file"""
        try:
            folder_path = os.path.dirname(filepath)
            if os.name == 'nt':  # Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # macOS, Linux
                if sys.platform == "darwin":
                    subprocess.call(('open', folder_path))
                else:
                    subprocess.call(('xdg-open', folder_path))
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder: {str(e)}")
    
    def set_mode(self, mode):
        """Set the window mode: 'add' or 'view'"""
        self.mode = mode
        
        if mode == "add":
            self.title_label.config(text="Add New Transaction")
            self.form_frame.pack(fill='x', pady=10)
            self.view_frame.pack(fill='both', expand=True, pady=10)
        elif mode == "view":
            self.title_label.config(text="View Transactions")
            self.form_frame.pack_forget()
            self.view_frame.pack(fill='both', expand=True, pady=10)
            self.load_transactions()
    
    def show(self):
        self.frame.pack(expand=True, fill='both')
        if self.mode == "view":
            self.load_transactions()
    
    def hide(self):
        self.frame.pack_forget()