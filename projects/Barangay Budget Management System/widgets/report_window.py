# widgets/report_window.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.page import PageMargins

class ReportWindow:
    def __init__(self, parent, auth_system, report_generator, on_back):
        self.parent = parent
        self.auth_system = auth_system
        self.report_generator = report_generator
        self.on_back = on_back
        
        self.frame = ttk.Frame(parent)
        self.report_type = "daily"  # "daily" or "monthly"
        self.create_widgets()
    
    def create_widgets(self):
        # Header
        header_frame = ttk.Frame(self.frame)
        header_frame.pack(fill='x', padx=20, pady=10)
        
        ttk.Button(header_frame, text="← Back to Dashboard", 
                  command=self.on_back).pack(side='left')
        
        self.title_label = ttk.Label(header_frame, text="Daily Report", 
                                    font=('Arial', 16, 'bold'))
        self.title_label.pack(side='left', padx=20)
        
        # Report type selector
        type_frame = ttk.Frame(header_frame)
        type_frame.pack(side='right')
        
        self.report_type_var = tk.StringVar(value="daily")
        ttk.Radiobutton(type_frame, text="Daily Report", variable=self.report_type_var,
                       value="daily", command=self.switch_report_type).pack(side='left', padx=5)
        ttk.Radiobutton(type_frame, text="Monthly Report", variable=self.report_type_var,
                       value="monthly", command=self.switch_report_type).pack(side='left', padx=5)
        
        # Main content
        self.main_frame = ttk.Frame(self.frame)
        self.main_frame.pack(expand=True, fill='both', padx=20, pady=10)
        
        self.create_daily_controls()
        self.create_monthly_controls()
        self.create_report_display()
    
    def create_daily_controls(self):
        """Create controls for daily report"""
        self.daily_controls = ttk.Frame(self.main_frame)
        
        ttk.Label(self.daily_controls, text="Select Date:").pack(side='left', padx=5)
        
        # Date entry
        self.daily_date = ttk.Entry(self.daily_controls, width=12)
        self.daily_date.insert(0, date.today().isoformat())
        self.daily_date.pack(side='left', padx=5)
        
        # Generate button
        ttk.Button(self.daily_controls, text="Generate Report",
                  command=self.generate_daily_report).pack(side='left', padx=10)
        
        # Export button for daily report
        ttk.Button(self.daily_controls, text="Export to Excel",
                  command=self.export_daily_excel).pack(side='left', padx=5)
    
    def create_monthly_controls(self):
        """Create controls for monthly report"""
        self.monthly_controls = ttk.Frame(self.main_frame)
        
        ttk.Label(self.monthly_controls, text="Select Month:").pack(side='left', padx=5)
        
        # Month selection
        current_year = datetime.now().year
        self.month_var = tk.StringVar(value=str(datetime.now().month))
        self.year_var = tk.StringVar(value=str(current_year))
        
        month_combo = ttk.Combobox(self.monthly_controls, textvariable=self.month_var,
                                  values=[str(i) for i in range(1, 13)], width=3,
                                  state='readonly')
        month_combo.pack(side='left', padx=5)
        
        ttk.Label(self.monthly_controls, text="/").pack(side='left')
        
        year_combo = ttk.Combobox(self.monthly_controls, textvariable=self.year_var,
                                 values=[str(i) for i in range(current_year-5, current_year+1)],
                                 width=5, state='readonly')
        year_combo.pack(side='left', padx=5)
        
        # Generate button
        ttk.Button(self.monthly_controls, text="Generate Report",
                  command=self.generate_monthly_report).pack(side='left', padx=10)
        
        # Export button for monthly report
        ttk.Button(self.monthly_controls, text="Export to Excel",
                  command=self.export_monthly_excel).pack(side='left', padx=5)
    
    def create_report_display(self):
        """Create the report display area"""
        # Summary frame
        self.summary_frame = ttk.LabelFrame(self.main_frame, text="Report Summary", padding="15")
        self.summary_frame.pack(fill='x', pady=10)
        
        # Create summary labels
        self.income_label = ttk.Label(self.summary_frame, text="Total Income: ₱0.00",
                                     font=('Arial', 11))
        self.income_label.grid(row=0, column=0, sticky='w', pady=5, padx=10)
        
        self.expense_label = ttk.Label(self.summary_frame, text="Total Expenses: ₱0.00",
                                      font=('Arial', 11))
        self.expense_label.grid(row=1, column=0, sticky='w', pady=5, padx=10)
        
        self.net_label = ttk.Label(self.summary_frame, text="Net Flow: ₱0.00",
                                  font=('Arial', 11, 'bold'))
        self.net_label.grid(row=2, column=0, sticky='w', pady=5, padx=10)
        
        self.count_label = ttk.Label(self.summary_frame, text="Total Transactions: 0",
                                    font=('Arial', 10))
        self.count_label.grid(row=0, column=1, sticky='w', pady=5, padx=10)
        
        self.period_label = ttk.Label(self.summary_frame, text="Period: ",
                                     font=('Arial', 10))
        self.period_label.grid(row=1, column=1, sticky='w', pady=5, padx=10)
        
        # Category breakdown frame
        self.category_frame = ttk.LabelFrame(self.main_frame, text="Category Breakdown", padding="10")
        self.category_frame.pack(fill='both', expand=True, pady=10)
        
        # Treeview for category breakdown
        columns = ('category', 'type', 'amount', 'percentage')
        self.category_tree = ttk.Treeview(self.category_frame, columns=columns, show='headings', height=12)
        
        # Define headings
        self.category_tree.heading('category', text='Category')
        self.category_tree.heading('type', text='Type')
        self.category_tree.heading('amount', text='Amount')
        self.category_tree.heading('percentage', text='Percentage')
        
        # Configure columns
        self.category_tree.column('category', width=200)
        self.category_tree.column('type', width=80)
        self.category_tree.column('amount', width=120)
        self.category_tree.column('percentage', width=100)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.category_frame, orient=tk.VERTICAL, command=self.category_tree.yview)
        self.category_tree.configure(yscrollcommand=scrollbar.set)
        
        self.category_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
    
    def switch_report_type(self):
        """Switch between daily and monthly reports"""
        self.report_type = self.report_type_var.get()
        
        if self.report_type == "daily":
            self.title_label.config(text="Daily Report")
            self.daily_controls.pack(fill='x', pady=10)
            self.monthly_controls.pack_forget()
            self.generate_daily_report()
        else:
            self.title_label.config(text="Monthly Report")
            self.monthly_controls.pack(fill='x', pady=10)
            self.daily_controls.pack_forget()
            self.generate_monthly_report()
    
    def generate_daily_report(self):
        """Generate and display daily report"""
        try:
            report_date = self.daily_date.get()
            report = self.report_generator.generate_daily_report(report_date)
            
            self.display_report(report)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate daily report: {str(e)}")
    
    def generate_monthly_report(self):
        """Generate and display monthly report"""
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
            report = self.report_generator.generate_monthly_report(year, month)
            
            self.display_report(report)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate monthly report: {str(e)}")
    
    def export_daily_excel(self):
        """Export daily report to Excel with landscape orientation"""
        try:
            report_date = self.daily_date.get()
            report = self.report_generator.generate_daily_report(report_date)
            self.export_to_excel(report, "daily")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export daily report: {str(e)}")
    
    def export_monthly_excel(self):
        """Export monthly report to Excel with landscape orientation"""
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
            report = self.report_generator.generate_monthly_report(year, month)
            self.export_to_excel(report, "monthly")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export monthly report: {str(e)}")
    
    def export_to_excel(self, report, report_type):
        """Export report data to Excel with landscape orientation"""
        # Ask user for save location
        default_filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Set landscape orientation and page setup
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Set margins
            ws.page_margins = PageMargins(
                left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
            )
            
            # Title
            title = f"{report_type.capitalize()} Financial Report"
            if report_type == "daily":
                title += f" - {report.get('date', 'Unknown Date')}"
            else:
                title += f" - {report.get('period', 'Unknown Period')}"
            
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')  # Updated to F1 for 6 columns
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Report generation info
            ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(italic=True)
            
            # Summary section
            ws['A5'] = "Summary"
            ws['A5'].font = Font(size=14, bold=True)
            
            summary_data = [
                ["Total Income:", f"₱{report['total_income']:,.2f}"],
                ["Total Expenses:", f"₱{report['total_expenses']:,.2f}"],
                ["Net Flow:", f"₱{report['net_flow']:,.2f}"],
                ["Total Transactions:", str(report['transaction_count'])]
            ]
            
            for i, (label, value) in enumerate(summary_data, start=6):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
            
            # Category Breakdown section
            ws['A10'] = "Category Breakdown"
            ws['A10'].font = Font(size=14, bold=True)
            
            # Headers for category breakdown
            headers = ['Category', 'Type', 'Amount', 'Percentage']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=11, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Category data
            row = 12
            total = report['total_income'] + report['total_expenses']
            
            # Get category breakdown from report
            categories = report.get('category_breakdown', [])
            if not categories and report.get('transactions'):
                # Group transactions by category manually if not provided
                categories = {}
                for trans in report['transactions']:
                    cat_name = trans['category_name']
                    if cat_name not in categories:
                        categories[cat_name] = {
                            'category_name': cat_name,
                            'category_type': trans['category_type'],
                            'total': 0
                        }
                    categories[cat_name]['total'] += trans['amount']
                categories = list(categories.values())
            
            for category in categories:
                percentage = (category['total'] / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=1).value = category['category_name']
                ws.cell(row=row, column=2).value = category['category_type'].capitalize()
                ws.cell(row=row, column=3).value = f"₱{category['total']:,.2f}"
                ws.cell(row=row, column=4).value = f"{percentage:.1f}%"
                row += 1
            
            # Transaction Details section (if available)
            if report.get('transactions'):
                row += 2
                ws.cell(row=row, column=1).value = "Transaction Details"
                ws.cell(row=row, column=1).font = Font(size=14, bold=True)
                row += 1
                
                # Headers for transaction details
                trans_headers = ['Transaction No.', 'Date', 'Description', 'Category', 'Type', 'Amount']
                for col, header in enumerate(trans_headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                
                row += 1
                # Add transactions with ACTUAL transaction numbers
                for trans in report['transactions']:
                    # Use the actual transaction_number from the transaction data
                    transaction_number = trans.get('transaction_number', 'N/A')
                    
                    # Get date directly from database using the correct field name
                    date_value = trans.get('transaction_date', '')  # UPDATED: changed from 'date' to 'transaction_date'
                    
                    ws.cell(row=row, column=1).value = transaction_number
                    ws.cell(row=row, column=2).value = date_value  # Direct from database
                    ws.cell(row=row, column=3).value = trans.get('description', '')
                    ws.cell(row=row, column=4).value = trans.get('category_name', '')
                    ws.cell(row=row, column=5).value = trans.get('category_type', '').capitalize()
                    ws.cell(row=row, column=6).value = f"₱{trans.get('amount', 0):,.2f}"
                    row += 1
            
            # Simple column width adjustment
            self.simple_adjust_column_widths(ws)
            
            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Report exported successfully to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel file: {str(e)}")

    def simple_adjust_column_widths(self, worksheet):
        """Simple column width adjustment that avoids merged cell issues"""
        # Updated column widths to include transaction number column
        column_widths = {
            'A': 15,  # Transaction Number (wider for actual transaction numbers)
            'B': 12,  # Date
            'C': 25,  # Description
            'D': 15,  # Category
            'E': 12,  # Type
            'F': 15,  # Amount
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
    
    def display_report(self, report):
        """Display the generated report"""
        # Update summary
        self.income_label.config(text=f"Total Income: ₱{report['total_income']:,.2f}")
        self.expense_label.config(text=f"Total Expenses: ₱{report['total_expenses']:,.2f}")
        self.net_label.config(text=f"Net Flow: ₱{report['net_flow']:,.2f}")
        
        # Color code net flow
        if report['net_flow'] > 0:
            self.net_label.config(foreground='green')
        elif report['net_flow'] < 0:
            self.net_label.config(foreground='red')
        else:
            self.net_label.config(foreground='black')
        
        # Update period and count
        if 'date' in report:
            self.period_label.config(text=f"Date: {report['date']}")
            self.count_label.config(text=f"Transactions: {report['transaction_count']}")
        else:
            self.period_label.config(text=f"Period: {report['period']}")
            self.count_label.config(text=f"Transactions: {report['transaction_count']}")
        
        # Update category breakdown
        self.update_category_tree(report)
    
    def update_category_tree(self, report):
        """Update the category breakdown treeview"""
        # Clear existing data
        for item in self.category_tree.get_children():
            self.category_tree.delete(item)
        
        # Calculate totals for percentage
        total = report['total_income'] + report['total_expenses']
        
        # Add categories to treeview
        for category in report.get('category_breakdown', []):
            percentage = (category['total'] / total * 100) if total > 0 else 0
            
            self.category_tree.insert('', 'end', values=(
                category['category_name'],
                category['category_type'].capitalize(),
                f"₱{category['total']:,.2f}",
                f"{percentage:.1f}%"
            ))
        
        # If no category breakdown, check for transactions
        if not report.get('category_breakdown') and report.get('transactions'):
            # Group transactions by category manually
            categories = {}
            for trans in report['transactions']:
                cat_name = trans['category_name']
                if cat_name not in categories:
                    categories[cat_name] = {
                        'category_name': cat_name,
                        'category_type': trans['category_type'],
                        'total': 0
                    }
                categories[cat_name]['total'] += trans['amount']
            
            for category in categories.values():
                percentage = (category['total'] / total * 100) if total > 0 else 0
                
                self.category_tree.insert('', 'end', values=(
                    category['category_name'],
                    category['category_type'].capitalize(),
                    f"₱{category['total']:,.2f}",
                    f"{percentage:.1f}%"
                ))
    
    def show(self):
        self.frame.pack(expand=True, fill='both')
        self.switch_report_type()  # Generate initial report
    
    def hide(self):
        self.frame.pack_forget()