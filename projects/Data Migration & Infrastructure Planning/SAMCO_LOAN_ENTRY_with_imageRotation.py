import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
import os
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

class SAMCOLoanEntry:
    def __init__(self, root):
        self.root = root
        self.root.title("SAMCO LOAN ENTRY SYSTEM")
        self.root.geometry("1400x900")
        
        # Configure styles
        style = ttk.Style()
        style.configure('Title.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Heading.TLabel', font=('Arial', 11, 'bold'))
        style.configure('Info.TLabel', font=('Arial', 9))
        style.configure('TLabelframe.Label', font=('Arial', 10, 'bold'))
        
        # Variables
        self.folder_path = None
        self.excel_path = None
        self.sheet_name = "ar_loan"          # FIXED: added sheet name
        self.image_files = []
        self.current_image_index = 0
        self.client_df = None
        self.progress_file = "loan_entry_progress.json"
        self.payment_mode = False
        
        # Image rotation variables
        self.original_images = {}  # Store original PIL images
        self.image_rotations = {}  # Store rotation angles per image
        self.current_photo = None   # Current PhotoImage reference
        
        # Load saved progress
        self.load_progress()
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=2)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # File selection frame (top)
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # Folder selection
        ttk.Label(file_frame, text="Loan Documents Folder:", font=('Arial', 10)).grid(row=0, column=0, sticky=tk.W)
        self.folder_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.folder_var, width=60, font=('Arial', 10)).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.select_folder, width=12).grid(row=0, column=2)
        
        # Excel file selection
        ttk.Label(file_frame, text="Excel File (with 'client' sheet):", font=('Arial', 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.excel_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.excel_var, width=60, font=('Arial', 10)).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.select_excel, width=12).grid(row=1, column=2)
        ttk.Button(file_frame, text="Load Data", command=self.load_data, width=12).grid(row=1, column=3, padx=5)
        
        # Image preview frame
        preview_frame = ttk.LabelFrame(main_frame, text="Document Preview (Reference Only)", padding="10")
        preview_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5, padx=(0, 5))
        
        # Create a container for canvas and rotation toolbar
        preview_container = ttk.Frame(preview_frame)
        preview_container.pack(fill=tk.BOTH, expand=True)
        
        # Rotation toolbar (TOP of preview area)
        rotation_toolbar = ttk.Frame(preview_container)
        rotation_toolbar.pack(side=tk.TOP, fill=tk.X, pady=(0, 5))
        
        ttk.Label(rotation_toolbar, text="Image Controls:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=5)
        ttk.Button(rotation_toolbar, text="↺ Rotate Left", command=self.rotate_left, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(rotation_toolbar, text="↻ Rotate Right", command=self.rotate_right, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(rotation_toolbar, text="⟳ Reset", command=self.reset_rotation, width=8).pack(side=tk.LEFT, padx=2)
        
        # Canvas for image display
        self.canvas = tk.Canvas(preview_container, bg='gray90', highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(preview_container, orient=tk.VERTICAL, command=self.canvas.yview)
        scrollbar_x = ttk.Scrollbar(preview_container, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.image_on_canvas = self.canvas.create_image(0, 0, anchor=tk.NW, image=None)
        
        # Navigation frame (BOTTOM of preview area)
        nav_frame = ttk.Frame(preview_frame)
        nav_frame.pack(side=tk.BOTTOM, pady=10)
        
        ttk.Button(nav_frame, text="◄ PREVIOUS", command=self.prev_image, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(nav_frame, text="NEXT ►", command=self.next_image, width=12).pack(side=tk.LEFT, padx=5)
        self.image_counter = ttk.Label(nav_frame, text="", font=('Arial', 10, 'bold'))
        self.image_counter.pack(side=tk.LEFT, padx=15)
        
        # Add zoom hint
        ttk.Label(nav_frame, text="(Image auto-fits to window)", font=('Arial', 8, 'italic'), foreground='gray50').pack(side=tk.LEFT, padx=10)
        
        # Loan entry form frame
        form_frame = ttk.LabelFrame(main_frame, text="Loan Entry Form", padding="10")
        form_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5, padx=(5, 0))
        
        form_canvas = tk.Canvas(form_frame, highlightthickness=0)
        form_scrollbar = ttk.Scrollbar(form_frame, orient=tk.VERTICAL, command=form_canvas.yview)
        form_canvas.configure(yscrollcommand=form_scrollbar.set)
        
        form_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        form_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        form_interior = ttk.Frame(form_canvas)
        form_canvas.create_window((0, 0), window=form_interior, anchor=tk.NW)
        
        row = 0
        
        # Borrower Information
        ttk.Label(form_interior, text="BORROWER INFORMATION", font=('Arial', 11, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        row += 1
        
        ttk.Label(form_interior, text="Last Name:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.lname_var = tk.StringVar()
        self.lname_entry = ttk.Entry(form_interior, textvariable=self.lname_var, width=25, font=('Arial', 10))
        self.lname_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="First Name:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.fname_var = tk.StringVar()
        self.fname_entry = ttk.Entry(form_interior, textvariable=self.fname_var, width=25, font=('Arial', 10))
        self.fname_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="Middle Name:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.mname_var = tk.StringVar()
        self.mname_entry = ttk.Entry(form_interior, textvariable=self.mname_var, width=25, font=('Arial', 10))
        self.mname_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Button(form_interior, text="🔍 SEARCH CLIENT ID", command=self.search_client, width=20).grid(row=row, column=0, columnspan=2, pady=10)
        row += 1
        
        ttk.Label(form_interior, text="Client ID:", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.client_id_var = tk.StringVar()
        self.client_id_display = ttk.Entry(form_interior, textvariable=self.client_id_var, width=25, font=('Arial', 10, 'bold'), state='readonly')
        self.client_id_display.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Separator(form_interior, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=15)
        row += 1
        
        # Loan Details
        ttk.Label(form_interior, text="LOAN DETAILS", font=('Arial', 11, 'bold')).grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        row += 1
        
        ttk.Label(form_interior, text="SLC_CODE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.slc_var = tk.StringVar()
        self.slc_combo = ttk.Combobox(form_interior, textvariable=self.slc_var, values=['12', '13', '14', '15'], width=23, state='readonly', font=('Arial', 10))
        self.slc_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        ttk.Label(form_interior, text="12-Loans Rec., 13-Accounts Rec., 14-Advances, 15-Other Rec.", font=('Arial', 8, 'italic'), foreground='gray30').grid(row=row, column=2, sticky=tk.W, padx=10)
        row += 1
        
        ttk.Label(form_interior, text="SLT_CODE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.slt_var = tk.StringVar()
        slt_values = [f"{i:03d}" for i in range(1, 8)]
        self.slt_combo = ttk.Combobox(form_interior, textvariable=self.slt_var, values=slt_values, width=23, state='readonly', font=('Arial', 10))
        self.slt_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        notes_text = "001-Arise, 002-Agri, 003-PO Fin, 004-Trade, 005-Non-Trade, 006-Employee, 007-Consultant"
        ttk.Label(form_interior, text=notes_text, font=('Arial', 8, 'italic'), foreground='gray30').grid(row=row, column=2, sticky=tk.W, padx=10)
        row += 1
        
        ttk.Label(form_interior, text="STATUS:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(form_interior, textvariable=self.status_var, values=['10 (PASTDUE)', '11 (CURRENT)'], width=23, state='readonly', font=('Arial', 10))
        self.status_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        ttk.Label(form_interior, text="10-Prior(PASTDUE), 11-Current(MIGZ)", font=('Arial', 8, 'italic'), foreground='gray30').grid(row=row, column=2, sticky=tk.W, padx=10)
        row += 1
        
        ttk.Label(form_interior, text="SLTYPE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.sltype_var = tk.StringVar()
        self.sltype_combo = ttk.Combobox(form_interior, textvariable=self.sltype_var, 
                                    values=['SALARY LOAN', 'SMALL BUSINESS LOAN', 'EMERGENCY LOAN', 'HONORARIUM LOAN'], 
                                    width=23, state='readonly', font=('Arial', 10))
        self.sltype_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="TR_CODE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.tr_var = tk.StringVar()
        self.tr_combo = ttk.Combobox(form_interior, textvariable=self.tr_var, values=['BEGIN', 'CR'], width=23, state='readonly', font=('Arial', 10))
        self.tr_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="TR_DATE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.tr_date_var = tk.StringVar()
        self.tr_date_var.set(datetime.now().strftime('%Y-%m-%d'))
        self.tr_date_entry = ttk.Entry(form_interior, textvariable=self.tr_date_var, width=25, font=('Arial', 10))
        self.tr_date_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="SETUPDATE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.setupdate_var = tk.StringVar()
        self.setupdate_var.set(datetime.now().strftime('%Y-%m-%d'))
        self.setupdate_entry = ttk.Entry(form_interior, textvariable=self.setupdate_var, width=25, font=('Arial', 10))
        self.setupdate_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="DEBIT (Loan Amount):", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.debit_var = tk.StringVar()
        self.debit_entry = ttk.Entry(form_interior, textvariable=self.debit_var, width=25, font=('Arial', 10))
        self.debit_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        # New fields: Interest, Principal, Service Fee, CBU Retention, Insurance
        ttk.Label(form_interior, text="INTEREST:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.interest_var = tk.StringVar()
        # Use trace_add for Python 3.14+ compatibility
        self.interest_var.trace_add('write', lambda *args: self.update_credit_total())
        self.interest_entry = ttk.Entry(form_interior, textvariable=self.interest_var, width=25, font=('Arial', 10))
        self.interest_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="PRINCIPAL:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.principal_var = tk.StringVar()
        self.principal_var.trace_add('write', lambda *args: self.update_credit_total())
        self.principal_entry = ttk.Entry(form_interior, textvariable=self.principal_var, width=25, font=('Arial', 10))
        self.principal_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="SERVICE FEE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.service_fee_var = tk.StringVar()
        self.service_fee_var.trace_add('write', lambda *args: self.update_credit_total())
        self.service_fee_entry = ttk.Entry(form_interior, textvariable=self.service_fee_var, width=25, font=('Arial', 10))
        self.service_fee_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="CBU RETENTION:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.cbu_var = tk.StringVar()
        self.cbu_var.trace_add('write', lambda *args: self.update_credit_total())
        self.cbu_entry = ttk.Entry(form_interior, textvariable=self.cbu_var, width=25, font=('Arial', 10))
        self.cbu_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="INSURANCE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.insurance_var = tk.StringVar()
        self.insurance_var.trace_add('write', lambda *args: self.update_credit_total())
        self.insurance_entry = ttk.Entry(form_interior, textvariable=self.insurance_var, width=25, font=('Arial', 10))
        self.insurance_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        # AUTO-CALCULATED CREDIT field
        ttk.Label(form_interior, text="CREDIT (Auto-calculated):", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.credit_var = tk.StringVar()
        self.credit_entry = ttk.Entry(form_interior, textvariable=self.credit_var, width=25, font=('Arial', 10, 'bold'), state='readonly')
        self.credit_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="P_RUNBALANCE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.runbalance_var = tk.StringVar()
        self.runbalance_entry = ttk.Entry(form_interior, textvariable=self.runbalance_var, width=25, font=('Arial', 10))
        self.runbalance_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="INT_PAID:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.int_paid_var = tk.StringVar()
        self.int_paid_entry = ttk.Entry(form_interior, textvariable=self.int_paid_var, width=25, font=('Arial', 10))
        self.int_paid_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="TERMS:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.terms_var = tk.StringVar()
        self.terms_entry = ttk.Entry(form_interior, textvariable=self.terms_var, width=25, font=('Arial', 10))
        self.terms_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="TERM_PERD:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.term_perd_var = tk.StringVar()
        self.term_perd_entry = ttk.Entry(form_interior, textvariable=self.term_perd_var, width=25, font=('Arial', 10))
        self.term_perd_entry.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="INT_RATE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.intrate_var = tk.StringVar()
        self.intrate_combo = ttk.Combobox(form_interior, textvariable=self.intrate_var, values=['3', '4'], width=23, state='readonly', font=('Arial', 10))
        self.intrate_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="PAY_MODE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.paymode_var = tk.StringVar()
        self.paymode_combo = ttk.Combobox(form_interior, textvariable=self.paymode_var, values=['MONTHLY', 'DAILY', 'YEARLY'], width=23, state='readonly', font=('Arial', 10))
        self.paymode_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        ttk.Label(form_interior, text="AMMORTIZATIONTYPE:", font=('Arial', 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        self.ammort_var = tk.StringVar()
        self.ammort_combo = ttk.Combobox(form_interior, textvariable=self.ammort_var, values=['STRAIGHT', 'ANNUITY'], width=23, state='readonly', font=('Arial', 10))
        self.ammort_combo.grid(row=row, column=1, sticky=tk.W, padx=5)
        row += 1
        
        # Action buttons
        button_frame = ttk.Frame(form_interior)
        button_frame.grid(row=row, column=0, columnspan=3, pady=20)
        ttk.Button(button_frame, text="💾 SAVE RECORD", command=self.save_record, width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="🗑️ CLEAR FORM", command=self.clear_form, width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="💰 PAYMENTS", command=self.payment_mode_toggle, width=18).pack(side=tk.LEFT, padx=5)
        
        def configure_scroll_region(event):
            form_canvas.configure(scrollregion=form_canvas.bbox("all"))
        form_interior.bind("<Configure>", configure_scroll_region)
        
        self.progress_label = ttk.Label(main_frame, text="Ready", font=('Arial', 10, 'bold'), foreground="blue")
        self.progress_label.grid(row=2, column=0, columnspan=2, pady=10)
        
    def update_credit_total(self):
        """Calculate CREDIT as sum of Interest, Principal, Service Fee, CBU Retention, Insurance"""
        try:
            interest = float(self.interest_var.get() or 0)
            principal = float(self.principal_var.get() or 0)
            service_fee = float(self.service_fee_var.get() or 0)
            cbu = float(self.cbu_var.get() or 0)
            insurance = float(self.insurance_var.get() or 0)
            
            total = interest + principal + service_fee + cbu + insurance
            self.credit_var.set(f"{total:.2f}")
        except ValueError:
            self.credit_var.set("0.00")
        
    def select_folder(self):
        self.folder_path = filedialog.askdirectory(title="Select folder containing loan document images")
        if self.folder_path:
            self.folder_var.set(self.folder_path)
            self.load_images()
            
    def select_excel(self):
        self.excel_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if self.excel_path:
            self.excel_var.set(self.excel_path)
            
    def load_images(self):
        if not self.folder_path:
            return
        image_extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.gif'}
        self.image_files = []
        self.original_images = {}  # Clear cache
        self.image_rotations = {}  # Clear rotation memory
        
        for file in os.listdir(self.folder_path):
            if Path(file).suffix.lower() in image_extensions:
                self.image_files.append(os.path.join(self.folder_path, file))
        self.image_files.sort()
        
        if self.image_files:
            # Pre-load original images
            for img_path in self.image_files:
                try:
                    self.original_images[img_path] = Image.open(img_path).copy()
                    self.image_rotations[img_path] = 0
                except Exception as e:
                    print(f"Error loading {img_path}: {e}")
            
            if hasattr(self, 'last_saved_image') and self.last_saved_image:
                try:
                    self.current_image_index = self.image_files.index(self.last_saved_image)
                except ValueError:
                    self.current_image_index = 0
            else:
                self.current_image_index = 0
            self.display_image()
            self.update_image_counter()
            self.progress_label.config(text=f"Loaded {len(self.image_files)} images. Currently on image {self.current_image_index + 1} of {len(self.image_files)}")
        else:
            messagebox.showwarning("No Images", "No image files found in the selected folder!")
    
    def get_rotated_image(self, img_path, angle):
        """Get rotated image from original, applying current rotation"""
        if img_path not in self.original_images:
            return None
        
        original = self.original_images[img_path]
        if angle == 0:
            return original.copy()
        elif angle == 90:
            return original.rotate(-90, expand=True)
        elif angle == 180:
            return original.rotate(-180, expand=True)
        elif angle == 270:
            return original.rotate(-270, expand=True)
        return original.copy()
    
    def display_image(self):
        if not self.image_files or self.current_image_index >= len(self.image_files):
            return
        try:
            img_path = self.image_files[self.current_image_index]
            rotation_angle = self.image_rotations.get(img_path, 0)
            
            # Get rotated image
            img = self.get_rotated_image(img_path, rotation_angle)
            if img is None:
                return
            
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            
            if canvas_width > 100 and canvas_height > 100:
                img_ratio = img.width / img.height
                canvas_ratio = canvas_width / canvas_height
                if img_ratio > canvas_ratio:
                    new_width = canvas_width - 20
                    new_height = int(new_width / img_ratio)
                else:
                    new_height = canvas_height - 20
                    new_width = int(new_height * img_ratio)
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            else:
                img.thumbnail((800, 600), Image.Resampling.LANCZOS)
            
            photo = ImageTk.PhotoImage(img)
            self.canvas.itemconfig(self.image_on_canvas, image=photo)
            self.current_photo = photo  # Keep reference to prevent garbage collection
            self.canvas.image = photo
            self.canvas.configure(scrollregion=self.canvas.bbox(tk.ALL))
            self.update_image_counter()
        except Exception as e:
            self.canvas.itemconfig(self.image_on_canvas, image=None)
            self.canvas.create_text(400, 300, text=f"Error loading image:\n{str(e)}", font=('Arial', 12), fill='red')
    
    def rotate_left(self):
        """Rotate current image 90 degrees counter-clockwise"""
        if not self.image_files:
            return
        img_path = self.image_files[self.current_image_index]
        current_rotation = self.image_rotations.get(img_path, 0)
        new_rotation = (current_rotation + 90) % 360
        self.image_rotations[img_path] = new_rotation
        self.display_image()
        self.progress_label.config(text=f"Image rotated {new_rotation}°")
    
    def rotate_right(self):
        """Rotate current image 90 degrees clockwise"""
        if not self.image_files:
            return
        img_path = self.image_files[self.current_image_index]
        current_rotation = self.image_rotations.get(img_path, 0)
        new_rotation = (current_rotation - 90) % 360
        # Ensure positive angle
        if new_rotation < 0:
            new_rotation += 360
        self.image_rotations[img_path] = new_rotation
        self.display_image()
        self.progress_label.config(text=f"Image rotated {new_rotation}°")
    
    def reset_rotation(self):
        """Reset current image rotation to 0"""
        if not self.image_files:
            return
        img_path = self.image_files[self.current_image_index]
        self.image_rotations[img_path] = 0
        self.display_image()
        self.progress_label.config(text="Image rotation reset")
            
    def update_image_counter(self):
        if self.image_files:
            self.image_counter.config(text=f"Image {self.current_image_index + 1} of {len(self.image_files)}")
        else:
            self.image_counter.config(text="No images")
            
    def prev_image(self):
        if self.image_files and self.current_image_index > 0:
            self.current_image_index -= 1
            self.display_image()
            
    def next_image(self):
        if self.image_files and self.current_image_index < len(self.image_files) - 1:
            self.current_image_index += 1
            self.display_image()
            
    def load_data(self):
        if not self.excel_path:
            messagebox.showwarning("No File", "Please select the Excel file first!")
            return
        try:
            self.client_df = pd.read_excel(self.excel_path, sheet_name='client', header=None)
            self.client_df.columns = ['CLIENTID', 'LNAME', 'FNAME', 'MNAME']
            messagebox.showinfo("Success", f"Loaded {len(self.client_df)} client records")
            self.progress_label.config(text=f"Loaded {len(self.client_df)} clients from Excel")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
            
    def search_client(self):
        if self.client_df is None:
            messagebox.showwarning("No Data", "Please load the Excel file first!")
            return
        lname = self.lname_var.get().strip().upper()
        fname = self.fname_var.get().strip().upper()
        mname = self.mname_var.get().strip().upper()
        if not lname and not fname:
            messagebox.showwarning("Missing Info", "Please enter at least Last Name or First Name!")
            return
        mask = pd.Series([True] * len(self.client_df))
        if lname:
            mask &= self.client_df['LNAME'].str.upper().str.contains(lname, na=False)
        if fname:
            mask &= self.client_df['FNAME'].str.upper().str.contains(fname, na=False)
        if mname:
            mask &= self.client_df['MNAME'].str.upper().str.contains(mname, na=False)
        results = self.client_df[mask]
        if len(results) == 0:
            messagebox.showwarning("Not Found", "No matching client found!")
            self.client_id_var.set("")
        elif len(results) == 1:
            client_id = results.iloc[0]['CLIENTID']
            self.client_id_var.set(client_id)
            messagebox.showinfo("Found", f"Client ID: {client_id}")
        else:
            choices = [f"{row['CLIENTID']} - {row['LNAME']}, {row['FNAME']} {row['MNAME']}" for _, row in results.iterrows()]
            choice_window = tk.Toplevel(self.root)
            choice_window.title("Select Client")
            choice_window.geometry("500x400")
            ttk.Label(choice_window, text="Multiple clients found. Please select one:", font=('Arial', 10)).pack(pady=10)
            listbox = tk.Listbox(choice_window, height=10, font=('Arial', 10))
            listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            for choice in choices:
                listbox.insert(tk.END, choice)
            def on_select():
                selection = listbox.curselection()
                if selection:
                    selected = choices[selection[0]]
                    client_id = selected.split(' - ')[0]
                    self.client_id_var.set(client_id)
                    choice_window.destroy()
            ttk.Button(choice_window, text="Select", command=on_select).pack(pady=10)
    
    def payment_mode_toggle(self):
        if not self.payment_mode:
            if not self.client_id_var.get():
                messagebox.showwarning("No Client", "Please load a client record first (search or enter loan details) before using PAYMENTS.")
                return
            # Disable non-payment fields
            self.lname_entry.config(state='disabled')
            self.fname_entry.config(state='disabled')
            self.mname_entry.config(state='disabled')
            self.slc_combo.config(state='disabled')
            self.slt_combo.config(state='disabled')
            self.status_combo.config(state='disabled')
            self.sltype_combo.config(state='disabled')
            self.tr_combo.config(state='disabled')
            self.setupdate_entry.config(state='disabled')
            self.debit_entry.config(state='disabled')
            self.interest_entry.config(state='disabled')
            self.principal_entry.config(state='disabled')
            self.service_fee_entry.config(state='disabled')
            self.cbu_entry.config(state='disabled')
            self.insurance_entry.config(state='disabled')
            self.terms_entry.config(state='disabled')
            self.term_perd_entry.config(state='disabled')
            self.intrate_combo.config(state='disabled')
            self.paymode_combo.config(state='disabled')
            self.ammort_combo.config(state='disabled')
            # Enable payment fields and TR_DATE
            self.credit_entry.config(state='normal')
            self.runbalance_entry.config(state='normal')
            self.int_paid_entry.config(state='normal')
            self.tr_date_entry.config(state='normal')
            # Clear payment fields, set TR_DATE to today
            self.credit_var.set("")
            self.runbalance_var.set("")
            self.int_paid_var.set("")
            self.tr_date_var.set(datetime.now().strftime('%Y-%m-%d'))
            self.payment_mode = True
            self.progress_label.config(text="PAYMENT MODE: Edit CREDIT, P_RUNBALANCE, INT_PAID, TR_DATE. Other fields locked.")
            messagebox.showinfo("Payment Mode", "You are now in PAYMENT mode.\n\nEditable: CREDIT, P_RUNBALANCE, INT_PAID, TR_DATE.\nClick SAVE RECORD to add a payment row.")
        else:
            # Exit payment mode
            self.lname_entry.config(state='normal')
            self.fname_entry.config(state='normal')
            self.mname_entry.config(state='normal')
            self.slc_combo.config(state='readonly')
            self.slt_combo.config(state='readonly')
            self.status_combo.config(state='readonly')
            self.sltype_combo.config(state='readonly')
            self.tr_combo.config(state='readonly')
            self.tr_date_entry.config(state='normal')
            self.setupdate_entry.config(state='normal')
            self.debit_entry.config(state='normal')
            self.interest_entry.config(state='normal')
            self.principal_entry.config(state='normal')
            self.service_fee_entry.config(state='normal')
            self.cbu_entry.config(state='normal')
            self.insurance_entry.config(state='normal')
            self.terms_entry.config(state='normal')
            self.term_perd_entry.config(state='normal')
            self.intrate_combo.config(state='readonly')
            self.paymode_combo.config(state='readonly')
            self.ammort_combo.config(state='readonly')
            self.credit_entry.config(state='readonly')
            self.runbalance_entry.config(state='normal')
            self.int_paid_entry.config(state='normal')
            self.payment_mode = False
            self.progress_label.config(text="Normal mode. All fields editable.")
    
    def save_record(self):
        if not self.excel_path:
            messagebox.showwarning("No File", "Please select an Excel file first!")
            return
        if not self.client_id_var.get():
            messagebox.showwarning("Missing Data", "Please search and select a Client ID first!")
            return
        if not self.slc_var.get():
            messagebox.showwarning("Missing Data", "Please select SLC_CODE!")
            return
        if not self.slt_var.get():
            messagebox.showwarning("Missing Data", "Please select SLT_CODE!")
            return
        if not self.status_var.get():
            messagebox.showwarning("Missing Data", "Please select STATUS!")
            return
        if not self.sltype_var.get():
            messagebox.showwarning("Missing Data", "Please select SLTYPE!")
            return
        if not self.tr_var.get():
            messagebox.showwarning("Missing Data", "Please select TR_CODE!")
            return
        if not self.intrate_var.get():
            messagebox.showwarning("Missing Data", "Please select INT_RATE!")
            return
        if not self.paymode_var.get():
            messagebox.showwarning("Missing Data", "Please select PAY_MODE!")
            return
        if not self.ammort_var.get():
            messagebox.showwarning("Missing Data", "Please select AMMORTIZATIONTYPE!")
            return
        if not self.tr_date_var.get():
            messagebox.showwarning("Missing Data", "Please enter TR_DATE!")
            return
        if not self.setupdate_var.get():
            messagebox.showwarning("Missing Data", "Please enter SETUPDATE!")
            return
        
        status_value = self.status_var.get().split(' ')[0]
        
        # Prepare row data for 27 columns (A to AA)
        row_data = [
            self.client_id_var.get(),      # A
            self.slc_var.get(),            # B
            self.slt_var.get(),            # C
            status_value,                  # D
            '',                            # E
            self.sltype_var.get(),         # F
            self.tr_var.get(),             # G
            self.tr_date_var.get(),        # H
            self.setupdate_var.get(),      # I
            self.debit_var.get(),          # J
            self.credit_var.get(),         # K (Now auto-calculated)
            self.runbalance_var.get(),     # L
            '',                            # M
            self.int_paid_var.get(),       # N
            self.interest_var.get(),       # O (Interest)
            self.principal_var.get(),      # P (Principal)
            self.service_fee_var.get(),    # Q (Service Fee)
            self.terms_var.get(),          # R
            self.cbu_var.get(),            # S (CBU Retention)
            self.term_perd_var.get(),      # T
            self.intrate_var.get(),        # U
            self.insurance_var.get(),      # V (Insurance)
            self.paymode_var.get(),        # W
            '',                            # X
            '',                            # Y
            '',                            # Z
            self.ammort_var.get(),         # AA
        ]
        
        try:
            # Load or create workbook
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
            else:
                wb = Workbook()
                # Remove default sheet if present
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
            
            # Get or create sheet 'ar_loan'
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)
                # Add header row (A to AA) with descriptive names
                headers = ['CLIENTID', 'SLC_CODE', 'SLT_CODE', 'STATUS', 'E', 'SLTYPE', 'TR_CODE', 'TR_DATE', 
                          'SETUPDATE', 'DEBIT', 'CREDIT', 'P_RUNBALANCE', 'M', 'INT_PAID', 'INTEREST', 
                          'PRINCIPAL', 'SERVICE_FEE', 'TERMS', 'CBU_RETENTION', 'TERM_PERD', 'INT_RATE', 
                          'INSURANCE', 'PAY_MODE', 'X', 'Y', 'Z', 'AMMORTIZATIONTYPE']
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
            
            next_row = ws.max_row + 1
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col_idx, value=value)
            
            wb.save(self.excel_path)
            
            messagebox.showinfo("Success", f"Record saved to {self.excel_path}\nSheet: {self.sheet_name}\nRow: {next_row}")
            
            # Save progress
            self.last_saved_image = self.image_files[self.current_image_index] if self.image_files else None
            self.save_progress()
            
            if self.payment_mode:
                # Clear payment fields for next payment, keep TR_DATE as today
                self.credit_var.set("")
                self.runbalance_var.set("")
                self.int_paid_var.set("")
                self.tr_date_var.set(datetime.now().strftime('%Y-%m-%d'))
                self.progress_label.config(text=f"Payment saved (row {next_row}). Ready for next payment.")
            else:
                self.clear_form()
                if self.image_files and self.current_image_index < len(self.image_files) - 1:
                    if messagebox.askyesno("Next Image", "Move to next image?"):
                        self.next_image()
                self.progress_label.config(text=f"Record saved successfully! Row {next_row} added to {self.sheet_name}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save record: {str(e)}\n\nMake sure the Excel file is not open in another program.")
            
    def clear_form(self):
        if self.payment_mode:
            self.payment_mode_toggle()
        self.lname_var.set("")
        self.fname_var.set("")
        self.mname_var.set("")
        self.client_id_var.set("")
        self.slc_var.set("")
        self.slt_var.set("")
        self.status_var.set("")
        self.sltype_var.set("")
        self.tr_var.set("")
        self.intrate_var.set("")
        self.paymode_var.set("")
        self.ammort_var.set("")
        self.debit_var.set("")
        self.interest_var.set("")
        self.principal_var.set("")
        self.service_fee_var.set("")
        self.cbu_var.set("")
        self.insurance_var.set("")
        self.credit_var.set("")
        self.runbalance_var.set("")
        self.int_paid_var.set("")
        self.terms_var.set("")
        self.term_perd_var.set("")
        self.tr_date_var.set(datetime.now().strftime('%Y-%m-%d'))
        self.setupdate_var.set(datetime.now().strftime('%Y-%m-%d'))
        
    def save_progress(self):
        progress_data = {
            'last_saved_image': self.last_saved_image if hasattr(self, 'last_saved_image') else None,
            'excel_path': self.excel_path,
            'folder_path': self.folder_path
        }
        try:
            with open(self.progress_file, 'w') as f:
                json.dump(progress_data, f)
        except:
            pass
            
    def load_progress(self):
        try:
            with open(self.progress_file, 'r') as f:
                progress_data = json.load(f)
                self.last_saved_image = progress_data.get('last_saved_image')
                if progress_data.get('excel_path') and os.path.exists(progress_data.get('excel_path')):
                    self.excel_path = progress_data.get('excel_path')
                    self.excel_var.set(self.excel_path)
                if progress_data.get('folder_path') and os.path.exists(progress_data.get('folder_path')):
                    self.folder_path = progress_data.get('folder_path')
                    self.folder_var.set(self.folder_path)
        except:
            self.last_saved_image = None

def main():
    root = tk.Tk()
    app = SAMCOLoanEntry(root)
    root.mainloop()

if __name__ == "__main__":
    main()