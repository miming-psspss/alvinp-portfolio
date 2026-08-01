"""
Batch Financial Report Processing for VAs, Part 2: Consolidation
Generic reference implementation for teaching purposes.

This tool demonstrates a repeatable pattern for:
  - Recursively finding and parsing structured .txt report files
  - Consolidating parsed transaction data into one master table
  - Exporting a formatted Excel (.xlsx) master sheet with a summary tab
  - Masking sensitive fields (account/card numbers) by default

Pairs well with Part 1: Extraction if your files start out inside
password-protected archives. This tool picks up from there, working
on the plain .txt files Part 1 produces, but also works standalone
on any similarly-formatted .txt files.

The file naming pattern to search for is set inside the app itself
(Configuration tab), not in this code, so no editing of this script
is ever required to use it with a new client.

Security note: sensitive-field masking is ON by default. This is a
deliberate choice, not an oversight: this tool handles real account
and card numbers, and a VA should have to actively choose to turn
masking off, with a clear warning, rather than accidentally export
or display unmasked financial identifiers.
"""

import os
import re
import calendar
import json
import threading
from datetime import datetime
from queue import Queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# =========================================================
# CONFIG: only change this if you want a different default
# starting value. The actual value used during a run always
# comes from the Configuration tab in the app, not from here.
# =========================================================
DEFAULT_TARGET_FILE_PREFIX = "REPORT_"
DEFAULT_TARGET_FILE_EXTENSIONS = ".txt"
APP_TITLE = "Batch Financial Report Processing for VAs, Part 2: Consolidation"
# =========================================================


def mask_value(value, keep_last=4):
    """Mask all but the last few characters of a sensitive value.
    '10000128' becomes 'XXXX0128'. Values too short to meaningfully
    mask are returned unchanged."""
    value = str(value)
    if len(value) <= keep_last:
        return value
    return ('X' * (len(value) - keep_last)) + value[-keep_last:]


class FinancialReportConsolidatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1400x1000")
        self.root.configure(bg='#f0f0f0')

        self.source_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.target_prefix = tk.StringVar(value=DEFAULT_TARGET_FILE_PREFIX)
        self.target_extensions = tk.StringVar(value=DEFAULT_TARGET_FILE_EXTENSIONS)
        self.mask_sensitive = tk.BooleanVar(value=True)  # ON by default, deliberately

        self.processing = False
        self.transactions_data = []
        self.total_files = 0
        self.total_transactions = 0
        self.log_queue = Queue()

        self.config_file = os.path.expanduser("~/financial_report_consolidator_part2_config.json")

        self.setup_styles()
        self.setup_gui()
        self.load_config()
        self.process_log_queue()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'), foreground='#34495e')
        style.configure('Success.TLabel', foreground='#27ae60')
        style.configure('Stats.TLabel', font=('Segoe UI', 11), foreground='#2c3e50')
        style.configure('Primary.TButton', font=('Segoe UI', 10, 'bold'))

    def setup_gui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(main_frame, text=APP_TITLE, style='Title.TLabel')
        title_label.pack(pady=(0, 20))

        if not OPENPYXL_AVAILABLE:
            warning = ttk.Label(
                main_frame,
                text="openpyxl is not installed. Run install_setup.bat first, then restart this tool.",
                foreground='red', font=('Segoe UI', 10, 'bold'))
            warning.pack(pady=(0, 10))

        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(paned, width=420)
        paned.add(left_frame, weight=1)
        self.setup_left_panel(left_frame)

        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=3)
        self.setup_right_panel(right_frame)

        self.setup_status_bar(main_frame)

    def setup_left_panel(self, parent):
        # Create scrollable container
        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Now use scrollable_frame as the parent for all content
        container = scrollable_frame
        
        # File Locations section
        path_frame = ttk.LabelFrame(container, text="File Locations", padding="15")
        path_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(path_frame, text="Source Folder (searched recursively):",
                style='Header.TLabel').pack(anchor=tk.W, pady=(0, 5))
        source_frame = ttk.Frame(path_frame)
        source_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Entry(source_frame, textvariable=self.source_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(source_frame, text="Browse", command=self.browse_source,
                style='Primary.TButton').pack(side=tk.RIGHT)
        
        ttk.Label(path_frame, text="Output Folder (for the Excel master sheet):",
                style='Header.TLabel').pack(anchor=tk.W, pady=(10, 5))
        output_frame = ttk.Frame(path_frame)
        output_frame.pack(fill=tk.X)
        ttk.Entry(output_frame, textvariable=self.output_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_frame, text="Browse", command=self.browse_output,
                style='Primary.TButton').pack(side=tk.RIGHT)
        
        # Target File Settings section
        target_frame = ttk.LabelFrame(container, text="Target File Settings", padding="15")
        target_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(target_frame, text="What do the files start with?",
                style='Header.TLabel').pack(anchor=tk.W)
        ttk.Entry(target_frame, textvariable=self.target_prefix).pack(fill=tk.X, pady=(5, 10))
        ttk.Label(target_frame, text=f"Example: {DEFAULT_TARGET_FILE_PREFIX}",
                font=('Segoe UI', 8, 'italic')).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Label(target_frame, text="File type(s), comma-separated if more than one:",
                style='Header.TLabel').pack(anchor=tk.W)
        ttk.Entry(target_frame, textvariable=self.target_extensions).pack(fill=tk.X, pady=(5, 5))
        ttk.Label(target_frame, text="Example: .txt or .txt, .csv",
                font=('Segoe UI', 8, 'italic')).pack(anchor=tk.W)
        
        ttk.Label(target_frame,
                text="Ask your client if you're not sure. Get this from them, don't guess.",
                foreground='#8a6d00', wraplength=350).pack(anchor=tk.W, pady=(8, 0))
        
        # Security Settings section
        security_frame = ttk.LabelFrame(container, text="Security Settings", padding="15")
        security_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Checkbutton(
            security_frame,
            text="Mask account and card numbers (recommended, on by default)",
            variable=self.mask_sensitive,
            command=self.on_mask_toggle
        ).pack(anchor=tk.W)
        
        ttk.Label(
            security_frame,
            text="When on, account and card numbers show only the last 4 digits, "
                "everywhere: on screen, in searches, and in the exported Excel file. "
                "This protects real financial data from accidental exposure, for "
                "example if a screen is shared or a screenshot is taken.",
            foreground='#2c5f8a', wraplength=350, justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(6, 0))
        
        # Actions section
        action_frame = ttk.LabelFrame(container, text="Actions", padding="15")
        action_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.load_btn = ttk.Button(action_frame, text="Load Report Files",
                                    command=self.load_files, style='Primary.TButton')
        self.load_btn.pack(fill=tk.X, pady=(0, 10))
        
        self.export_btn = ttk.Button(action_frame, text="Generate Master Excel Sheet",
                                    command=self.generate_master_excel, style='Primary.TButton')
        self.export_btn.pack(fill=tk.X, pady=(0, 10))
        self.export_btn.config(state='disabled')
        
        ttk.Button(action_frame, text="Clear All Data",
                command=self.clear_data).pack(fill=tk.X)
        
        # Progress section
        progress_frame = ttk.LabelFrame(container, text="Progress", padding="15")
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        self.status_label = ttk.Label(progress_frame, text="Ready", font=('Segoe UI', 9))
        self.status_label.pack()
        
        # Statistics section
        stats_frame = ttk.LabelFrame(container, text="Statistics", padding="15")
        stats_frame.pack(fill=tk.X)
        self.files_var = tk.StringVar(value="Files: 0")
        self.trans_var = tk.StringVar(value="Transactions: 0")
        ttk.Label(stats_frame, textvariable=self.files_var, style='Stats.TLabel').pack(anchor=tk.W, pady=2)
        ttk.Label(stats_frame, textvariable=self.trans_var, style='Stats.TLabel').pack(anchor=tk.W, pady=2)


    def setup_right_panel(self, parent):
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(search_frame, text="Search:", style='Header.TLabel').pack(side=tk.LEFT, padx=(0, 10))
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', self.filter_treeview)
        ttk.Entry(search_frame, textvariable=self.search_var, width=30).pack(
            side=tk.LEFT, fill=tk.X, expand=True)

        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        columns = ('Year', 'Month', 'File', 'Date/Time', 'Seq No', 'Trace No',
                   'Tran Code', 'Account No', 'Card No', 'Amount', 'Fee', 'Terminal')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                  yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        y_scroll.config(command=self.tree.yview)
        x_scroll.config(command=self.tree.xview)

        column_widths = {
            'Year': 60, 'Month': 80, 'File': 150, 'Date/Time': 140,
            'Seq No': 80, 'Trace No': 80, 'Tran Code': 80,
            'Account No': 120, 'Card No': 120, 'Amount': 100,
            'Fee': 80, 'Terminal': 100
        }
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c))
            self.tree.column(col, width=column_widths.get(col, 100), minwidth=50)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind('<Double-Button-1>', self.show_transaction_details)

    def setup_status_bar(self, parent):
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        self.status_text = tk.StringVar(value="Ready. Select a source folder to begin.")
        ttk.Label(status_frame, textvariable=self.status_text, relief=tk.SUNKEN,
                  anchor=tk.W, padding=(10, 5)).pack(fill=tk.X)

    def update_status(self, message):
        self.status_text.set(message)

    def on_mask_toggle(self):
        """If the VA tries to turn masking off, make them confirm explicitly."""
        if not self.mask_sensitive.get():
            confirm = messagebox.askyesno(
                "Turn Off Masking?",
                "Turning this off will show and export FULL, unmasked account and "
                "card numbers.\n\n"
                "Only do this if your client has specifically asked for full numbers.\n\n"
                "Are you sure you want to turn off masking?"
            )
            if not confirm:
                self.mask_sensitive.set(True)
        if self.transactions_data:
            self.update_treeview()

    def browse_source(self):
        folder = filedialog.askdirectory(title="Select Folder with Report Files")
        if folder:
            self.source_path.set(folder)
            if not self.output_path.get():
                self.output_path.set(os.path.join(folder, "MASTER_SHEETS"))
            self.update_status(f"Source folder set: {folder}")

    def browse_output(self):
        folder = filedialog.askdirectory(title="Select Output Folder for the Excel file")
        if folder:
            self.output_path.set(folder)

    def load_files(self):
        source = self.source_path.get()
        if not source or not os.path.exists(source):
            messagebox.showerror("Error", "Please select a valid source folder!")
            return

        self.load_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.progress_bar.start()

        thread = threading.Thread(target=self.process_files, args=(source,))
        thread.daemon = True
        thread.start()

    def process_files(self, base_path):
        try:
            self.update_status("Scanning for report files...")
            self.transactions_data = []
            self.total_files = 0
            self.total_transactions = 0

            prefix = self.target_prefix.get().strip().upper()
            raw_extensions = self.target_extensions.get().strip()
            extensions = tuple(
                e.strip().lower() if e.strip().startswith('.') else '.' + e.strip().lower()
                for e in raw_extensions.split(',') if e.strip()
            )
            if not extensions:
                extensions = ('',)

            matched_files = []
            for root_dir, dirs, files in os.walk(base_path):
                for file in files:
                    if file.upper().startswith(prefix) and file.lower().endswith(extensions):
                        matched_files.append(os.path.join(root_dir, file))

            if not matched_files:
                self.root.after(0, lambda: messagebox.showinfo(
                    "Info", "No matching files found in the selected folder or its subfolders. "
                            "Check the Target File Settings match your client's file naming."))
                return

            for file_path in matched_files:
                self.process_single_file(file_path)
                self.total_files += 1
                self.root.after(0, lambda: self.files_var.set(f"Files: {self.total_files}"))
                self.root.after(0, lambda: self.trans_var.set(f"Transactions: {self.total_transactions}"))

            self.root.after(0, self.update_treeview)
            self.root.after(0, lambda: self.update_status(
                f"Loaded {self.total_files} files, {self.total_transactions} transactions"))
            if self.total_files > 0:
                self.root.after(0, lambda: self.export_btn.config(
                    state='normal' if OPENPYXL_AVAILABLE else 'disabled'))

        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.root.after(0, lambda: self.load_btn.config(state='normal'))
            self.root.after(0, lambda: self.progress_bar.stop())

    def process_single_file(self, file_path):
        """Parse a single report file. Metadata lines carry forward to the
        transaction lines that follow them, and Year/Month are derived from
        the report's own date, not from any folder naming convention."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                lines = content.split('\n')

            current_report_date = None
            current_run_time = None
            current_site_code = None
            current_terminal_id = None
            current_terminal_name = None
            year = "Unknown"
            month = "Unknown"

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                if 'FOR ' in line and not current_report_date:
                    date_match = re.search(r'FOR\s+(\d{1,2}/\d{1,2}/\d{4})', line)
                    if date_match:
                        current_report_date = date_match.group(1)
                        try:
                            mm, dd, yyyy = current_report_date.split('/')
                            year = yyyy
                            month = calendar.month_name[int(mm)]
                        except Exception:
                            pass

                if 'RUN TIME' in line and not current_run_time:
                    time_match = re.search(r'RUN TIME\s+(\d{2}:\d{2}:\d{2})', line)
                    if time_match:
                        current_run_time = time_match.group(1)

                if 'SITE CODE' in line and 'TERMINAL' not in line:
                    parts = line.split()
                    for part in parts:
                        if part.isdigit() and len(part) >= 4:
                            current_site_code = part
                            break

                if 'TERMINAL' in line:
                    parts = line.split()
                    for j, part in enumerate(parts):
                        if part == 'TERMINAL' and j + 1 < len(parts):
                            current_terminal_id = parts[j + 1]
                            if j + 2 < len(parts):
                                current_terminal_name = ' '.join(parts[j + 2:])
                            break

                if re.match(r'\s*\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}', line):
                    parts = re.split(r'\s{2,}', line)
                    if len(parts) >= 9:
                        transaction = {
                            'Year': year,
                            'Month': month,
                            'File': os.path.basename(file_path),
                            'ReportDate': current_report_date or '',
                            'RunTime': current_run_time or '',
                            'SiteCode': current_site_code or '',
                            'TerminalID': current_terminal_id or '',
                            'TerminalName': current_terminal_name or '',
                            'DateTime': parts[0],
                            'SeqNo': parts[1] if len(parts) > 1 else '',
                            'TraceNo': parts[2] if len(parts) > 2 else '',
                            'ATMTrace': parts[3] if len(parts) > 3 else '',
                            'TranCode': parts[4] if len(parts) > 4 else '',
                            'AccountNo': parts[5] if len(parts) > 5 else '',
                            'CardNo': parts[6] if len(parts) > 6 else '',
                            'TranAmount': parts[7].replace(',', '') if len(parts) > 7 else '',
                            'TranFee': parts[8].replace(',', '') if len(parts) > 8 else ''
                        }
                        self.transactions_data.append(transaction)
                        self.total_transactions += 1

        except Exception as e:
            self.log_queue.put((f"Error processing {os.path.basename(file_path)}: {str(e)}", "ERROR"))

    def display_value(self, trans, field):
        """Return the value for a field, masked if it's sensitive and masking is on."""
        value = trans.get(field, '')
        if field in ('AccountNo', 'CardNo') and self.mask_sensitive.get():
            return mask_value(value)
        return value

    def get_transaction_values(self, trans):
        return (
            trans['Year'], trans['Month'], trans['File'], trans['DateTime'],
            trans['SeqNo'], trans['TraceNo'], trans['TranCode'],
            self.display_value(trans, 'AccountNo'),
            self.display_value(trans, 'CardNo'),
            trans['TranAmount'], trans['TranFee'], trans['TerminalName']
        )

    def update_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for trans in self.transactions_data:
            self.tree.insert('', 'end', values=self.get_transaction_values(trans))
        self.files_var.set(f"Files: {self.total_files}")
        self.trans_var.set(f"Transactions: {self.total_transactions}")

    def filter_treeview(self, *args):
        search_term = self.search_var.get().lower()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for trans in self.transactions_data:
            values = self.get_transaction_values(trans)
            if not search_term or any(search_term in str(v).lower() for v in values):
                self.tree.insert('', 'end', values=values)

    def sort_treeview(self, col):
        items = [(self.tree.set(item, col), item) for item in self.tree.get_children('')]
        items.sort()
        for index, (val, item) in enumerate(items):
            self.tree.move(item, '', index)

    def show_transaction_details(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        item = self.tree.item(selection[0])
        values = item['values']
        columns = ['Year', 'Month', 'File', 'Date/Time', 'Seq No', 'Trace No',
                   'Tran Code', 'Account No', 'Card No', 'Amount', 'Fee', 'Terminal']

        popup = tk.Toplevel(self.root)
        popup.title("Transaction Details")
        popup.geometry("500x400")
        popup.resizable(False, False)
        frame = ttk.Frame(popup, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Transaction Details", style='Title.TLabel').pack(pady=(0, 20))

        if self.mask_sensitive.get():
            ttk.Label(frame, text="Sensitive fields are masked (Security Settings is ON).",
                      foreground='#2c5f8a').pack(pady=(0, 10))

        for i, col in enumerate(columns):
            row_frame = ttk.Frame(frame)
            row_frame.pack(fill=tk.X, pady=2)
            ttk.Label(row_frame, text=f"{col}:", font=('Segoe UI', 10, 'bold')).pack(
                side=tk.LEFT, padx=(0, 10))
            ttk.Label(row_frame, text=str(values[i]), font=('Segoe UI', 10)).pack(side=tk.LEFT)

        ttk.Button(frame, text="Close", command=popup.destroy).pack(pady=20)

    def generate_master_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is not installed. Please run install_setup.bat first, "
                "then restart this tool."
            )
            return

        if not self.transactions_data:
            messagebox.showwarning("Warning", "No data to export!")
            return

        output_dir = self.output_path.get()
        if not output_dir:
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Master Excel Sheet As"
            )
            if not output_file:
                return
        else:
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(output_dir, f"MASTER_TRANSACTIONS_{timestamp}.xlsx")

        try:
            wb = Workbook()
            self._write_master_sheet(wb.active)
            self._write_summary_sheet(wb.create_sheet("Summary"))
            wb.save(output_file)

            mask_note = "Sensitive fields were masked in this export." if self.mask_sensitive.get() \
                else "Sensitive fields were NOT masked in this export (masking was turned off)."
            self.update_status(f"Master Excel sheet generated: {output_file}")
            messagebox.showinfo(
                "Success",
                f"Master Excel sheet created successfully!\n\n"
                f"File: {output_file}\n"
                f"Total transactions: {len(self.transactions_data)}\n\n"
                f"{mask_note}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel file: {str(e)}")

    def _write_master_sheet(self, ws):
        ws.title = "Master Transactions"
        headers = [
            'Year', 'Month', 'Source File', 'Report Date', 'Run Time',
            'Site Code', 'Terminal ID', 'Terminal Name', 'Date/Time',
            'Seq No', 'Trace No', 'ATM Trace', 'Tran Code',
            'Account Number', 'Card Number', 'Transaction Amount', 'Transaction Fee'
        ]
        header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')

        for row_idx, trans in enumerate(self.transactions_data, start=2):
            account_no = self.display_value(trans, 'AccountNo')
            card_no = self.display_value(trans, 'CardNo')
            row_values = [
                trans['Year'], trans['Month'], trans['File'], trans['ReportDate'],
                trans['RunTime'], trans['SiteCode'], trans['TerminalID'], trans['TerminalName'],
                trans['DateTime'], trans['SeqNo'], trans['TraceNo'], trans['ATMTrace'],
                trans['TranCode'], account_no, card_no, trans['TranAmount'], trans['TranFee']
            ]
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        last_row = len(self.transactions_data) + 1
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        ws.freeze_panes = "A2"

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    def _write_summary_sheet(self, ws):
        ws.title = "Summary"
        header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ['Year', 'Month', 'Transaction Count', 'Total Amount', 'Total Fee']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        summary = {}
        for trans in self.transactions_data:
            key = (trans['Year'], trans['Month'])
            if key not in summary:
                summary[key] = {'count': 0, 'amount': 0.0, 'fee': 0.0}
            summary[key]['count'] += 1
            try:
                summary[key]['amount'] += float(trans['TranAmount']) if trans['TranAmount'] else 0.0
            except ValueError:
                pass
            try:
                summary[key]['fee'] += float(trans['TranFee']) if trans['TranFee'] else 0.0
            except ValueError:
                pass

        row_idx = 2
        for (year, month), data in sorted(summary.items()):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=month)
            ws.cell(row=row_idx, column=3, value=data['count'])
            ws.cell(row=row_idx, column=4, value=round(data['amount'], 2))
            ws.cell(row=row_idx, column=5, value=round(data['fee'], 2))
            row_idx += 1

        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 16)

    def clear_data(self):
        if messagebox.askyesno("Confirm", "Clear all loaded data?"):
            self.transactions_data = []
            self.total_files = 0
            self.total_transactions = 0
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.files_var.set("Files: 0")
            self.trans_var.set("Transactions: 0")
            self.export_btn.config(state='disabled')
            self.update_status("Data cleared")

    def process_log_queue(self):
        try:
            while not self.log_queue.empty():
                message, level = self.log_queue.get_nowait()
                print(f"[{level}] {message}")
        except Exception:
            pass
        finally:
            self.root.after(200, self.process_log_queue)

    def save_config(self):
        config = {
            'source_path': self.source_path.get(),
            'output_path': self.output_path.get(),
            'target_prefix': self.target_prefix.get(),
            'target_extensions': self.target_extensions.get(),
            'mask_sensitive': self.mask_sensitive.get(),
        }
        try:
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=4)
        except Exception:
            pass

    def load_config(self):
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                if config.get('source_path'):
                    self.source_path.set(config['source_path'])
                if config.get('output_path'):
                    self.output_path.set(config['output_path'])
                if config.get('target_prefix'):
                    self.target_prefix.set(config['target_prefix'])
                if config.get('target_extensions'):
                    self.target_extensions.set(config['target_extensions'])
                if 'mask_sensitive' in config:
                    self.mask_sensitive.set(config['mask_sensitive'])
        except Exception:
            pass

    def on_closing(self):
        self.save_config()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = FinancialReportConsolidatorGUI(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
