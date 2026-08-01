"""
Flexible case-data import.

Unlike a fixed-column import, this reads whatever sheet/header row the user
points at, and lets them confirm (or fix) how each Excel column maps to a
case field -- so it isn't tied to the original DATABASE sheet's exact layout.

Workflow:
    1. list_sheets(path)                          -> pick a sheet
    2. read_preview(path, sheet, max_rows=15)      -> show raw rows so the
                                                       user can spot the header row
    3. read_headers(path, sheet, header_row)       -> raw header strings
    4. guess_mapping(headers)                      -> best-guess field per column
    5. import_with_mapping(path, sheet, header_row, mapping) -> writes to cases.db
"""
import json
from pathlib import Path

import openpyxl

from app.db import get_connection, init_db
from app.schema_fields import guess_field_for_header, REQUIRED_FIELDS

MAPPINGS_FILE = Path(__file__).resolve().parent.parent / "data" / "import_mappings.json"


def list_sheets(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return wb.sheetnames


def read_preview(path: str, sheet: str, max_rows: int = 15):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        rows.append(row)
    return rows


def read_headers(path: str, sheet: str, header_row: int):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), start=header_row):
        return list(row)
    return []


def guess_mapping(headers: list) -> dict:
    """Returns {column_index (0-based): field_key or None}."""
    return {i: guess_field_for_header(h) for i, h in enumerate(headers)}


def _mapping_signature(headers: list) -> str:
    return "|".join(str(h).strip().lower() if h else "" for h in headers)


def load_saved_mapping(headers: list):
    if not MAPPINGS_FILE.exists():
        return None
    saved = json.loads(MAPPINGS_FILE.read_text())
    sig = _mapping_signature(headers)
    entry = saved.get(sig)
    if entry:
        return {int(k): v for k, v in entry.items()}
    return None


def save_mapping(headers: list, mapping: dict):
    MAPPINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    saved = {}
    if MAPPINGS_FILE.exists():
        saved = json.loads(MAPPINGS_FILE.read_text())
    sig = _mapping_signature(headers)
    saved[sig] = {str(k): v for k, v in mapping.items() if v}
    MAPPINGS_FILE.write_text(json.dumps(saved, indent=2))


def import_with_mapping(path: str, sheet: str, header_row: int, mapping: dict) -> int:
    """
    mapping: {column_index (0-based): field_key or None}
    Rows below header_row are imported; unmapped columns are ignored.
    """
    mapped_fields = [f for f in mapping.values() if f]
    missing_required = [f for f in REQUIRED_FIELDS if f not in mapped_fields]
    if missing_required:
        raise ValueError(
            f"These required fields aren't mapped to any column: {', '.join(missing_required)}. "
            "Every case needs at least these to be usable."
        )

    init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    col_to_field = {col: field for col, field in mapping.items() if field}
    fields = list(col_to_field.values())
    placeholders = ", ".join("?" for _ in fields)
    insert_sql = f"INSERT OR REPLACE INTO cases ({', '.join(fields)}) VALUES ({placeholders})"

    conn = get_connection()
    imported = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        # first mapped required field (case_no) must be present to count as a real row
        case_no_col = next((c for c, f in col_to_field.items() if f == "case_no"), None)
        if case_no_col is not None and (case_no_col >= len(row) or row[case_no_col] in (None, "")):
            continue
        values = []
        for col in col_to_field:
            val = row[col] if col < len(row) else None
            if hasattr(val, "date"):
                val = val.date().isoformat()
            values.append(val)
        conn.execute(insert_sql, values)
        imported += 1

    conn.commit()
    conn.close()

    if imported == 0:
        raise ValueError(
            "No case rows were found below the selected header row. Double-check the "
            "header row number is correct and that data actually starts right below it."
        )
    return imported
